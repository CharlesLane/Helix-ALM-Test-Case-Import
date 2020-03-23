from openpyxl import load_workbook
import os

# Load your workbook and worksheet
wb = load_workbook(r'C:\testfile.xlsx')
ws = wb.active

# Define your columns
tc = 1 # tc = test case name column
sc = 2 # sc = steps column
ec = 3 # ec = expected result column
cc = 4 # cc = comment column
ac = 5 # ac = additional column(s) starting point

# Define your starting row
cr = 2 # cr = current row

# Create text file
f = open('ConvertedTC.txt', 'w+')

# Count how many additional columns there are, and define them as numberOfColumns
numberOfColumns = 0
while ws.cell(row=1, column=ac).value != None:
    ac = ac + 1
    numberOfColumns += 1
ac = 5

# Write column headers to text file
for i in range(numberOfColumns):
    f.write(ws.cell(row=1, column=ac).value + ',')
    ac = ac + 1
f.write("Test Case,Steps\n")
ac = 5

# Loop through current Test Case Name
while cr >= 2:
    tcn = ws.cell(row=cr, column=tc).value # Define Test Case Name
    if ws.cell(row=cr, column=tc).value != None:
        # Add in any extra additional fields
        for i in range(numberOfColumns):
            f.write(str(ws.cell(row=cr, column=ac).value) + ',')
            ac = ac + 1
        f.write(str(tcn) + ',' + '"')
        ac = 5
    else:
        pass

    # Loop through steps and expected results, and write them to text file
    while tcn == ws.cell(row=cr, column=tc).value:
        if ws.cell(row=cr, column=sc).value != None:
            f.write('*' + str(ws.cell(row=cr, column=sc).value) + '\n')
        else:
            pass
        if ws.cell(row=cr, column=ec).value != None:
            f.write('+' + str(ws.cell(row=cr, column=ec).value) + '\n')
        else:
            pass
        if ws.cell(row=cr, column=cc).value != None:
            f.write('#' + str(ws.cell(row=cr, column=cc).value) + '\n') # Could add some logic in to differentiate between comments and step notes
        else:
            pass
        cr = cr + 1
        if (ws.cell(row=cr, column=sc).value == None and ws.cell(row=cr, column=ec).value == None):
            break
    f.write('"\n')
    if ws.cell(row=cr, column=tc).value == None:
        break
# Close the text file
f.close()

# Rename the text file as CSV
os.rename(r'ConvertedTC.txt',r'ConvertedTC.csv')